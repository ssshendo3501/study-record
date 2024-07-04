# 小学校入学年度
# 早生まれの対応の表を作る必要がある


import openpyxl as excel
import datetime 

book = excel.Workbook()
sheet = book.active

# 開始年の設定
base_year = datetime.date.today().year - 10

# 連続でｓ流に値を設定する
for i in range(80):
    year = base_year + i
    s1 = year - 7
    s2 = year - 6
    sf = f"{s1}年4月2日〜{s2}年4月1日に生まれた人"    
    # セルを取得し値を設定
    sheet.cell(i+1, 1, value=str(year)+"年度")
    sheet.cell(i+1, 2, value=sf)
    
book.save("nyugakunen_list.xlsx")