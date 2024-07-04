# 顧客リストから横浜と名古屋の顧客のみを抜き出す

import openpyxl as excel

# 顧客一覧のブックを開き名簿のシートを選択
book = excel.load_workbook("all-customer.xlsx")
sheet = book["名簿"]

# 抜き出す単語を記録する変数
customers = [["名前", "住所", "購入プラン"]]

# 顧客一覧を抽出
for row in sheet.iter_rows(min_row=3):
    values = [cell.value for cell in row]
    if values[0] is None: break
    # 横浜と名古屋ならコピー
    area = values[1]
    if area == "横浜市" or area == "名古屋市":
        customers.append(values)
        print(values)
        
# 新規ブックの作成
new_book = excel.Workbook()
new_sheet = new_book.active
new_sheet["A1"] = "横浜と名古屋の顧客名簿"
# 抽出したデータを繰り返しシートに書き込み
for row, row_val in enumerate(customers):
    for col, val in enumerate(row_val):
        c = new_sheet.cell(2+row, 1+col)
        c.value = val

new_book.save("yokohama_nagoya.xlsx")
        