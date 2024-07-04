# 100枚のシートを作る
# 100枚のシートのうち、1枚が当たり、99枚がハズレ

import openpyxl as excel
import random

# 当たりシートの番号を決める
atari = random.randint(1, 100)

# 新規ブックの作成
book = excel.Workbook()
book.active['B2'] = '当たりが書いたシートを探そう'

# 繰り返し100回シートを作成する
for i in range(1, 101):
    # 新規シート作成
    sheet_name = str(i) + '番'
    sheet = book.create_sheet(title=sheet_name)
    # シートに書き込む単語の決定
    word = 'ハズレ'
    if i == atari: word = '当たり'
    # インパクトがあるようにwordで画面を埋める
    for y in range(50):
        for x in range(30):
            c = sheet.cell(row=y+1, column=x+1)
            c.value = word

book.save("game100.xlsx")
print("ok, atari=", atari)


