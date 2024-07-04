import openpyxl as excel 

book = excel.load_workbook("uriage.xlsx")
sheet = book.active

# A3からF9のセルを取り出す
rows = sheet['A3':'F9']
for row in rows:
    values = [cell.value for cell in row]
    # # リスト内包表記で書かない場合
    # values = []
    # for cell in row:
    #     values.append(cell.value)
    print(values)


print('------------------------')
  
# 計算式ではなく計算済みの値が欲しい場合
# 引数にdata_only=Trueを加える
book = excel.load_workbook("uriage.xlsx", data_only=True)
sheet = book.active

rows = sheet['A3':'F9']
for row in rows:
    values = [cell.value for cell in row]
    print(values)
    
print('------------------------')

# データが何行あるか不明な場合
# 1. 適当な範囲まで処理
book = excel.load_workbook("uriage.xlsx", data_only=True)
sheet = book.active

# A3からF999まで抜き出し
rows = sheet['A3':'F999']
for row in rows:
    # セルの値をリストとして得る
    values = [cell.value for cell in row]
    # 空白セルであれば読み取りを終える
    if values[0] is None: break
    # リストを表示する
    print(values)
        
print('------------------------')
       
# 2. iter_rowsを使う
for row in sheet.iter_rows(min_row=3):
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values)
    
