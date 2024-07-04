import openpyxl as excel

book = excel.load_workbook("test100.xlsx")
sheet = book.active

# H2のセルを読み込む
print(sheet["H2"].value)

# H2のセルを読み込む（行と列を指定する方法）
cell = sheet.cell(row=2, column=8)
print(cell.value)

# for文を使って連続で値を読み込む
for y in range(2, 5):
    r = []
    for x in range(2, 5):
        v = sheet.cell(row=y, column=x).value
        r.append(v)
    print(r)
    
# ワークシートの特定の範囲を指定して値を読み込む
for row in sheet["B2":"D4"]:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)

# 上記のコードをリスト内包表記で書く
for row in sheet["B2":"D4"]:
    print([c.value for c in row])
    
    
#　イテレータ(iter_rows)で繰り返し指定範囲を得る方法
# 行番号/列番号を指定してイテレータを取得する
it = sheet.iter_rows(min_row=2, min_col=2,
                     max_row=4, max_col=4)

# for文でイテレータを回す
for row in it:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)
    
# セル名と行番号・列番号を変換する
cell = sheet["C2"]
(row, col) = (cell.row, cell.column)
print(f'C2=({row}, {col})')

# 行番号・列番号からセル名を得る
cell = sheet.cell(row=2, column=3)
cdt = cell.coordinate
print(f'cell(2, 3)={cdt}')