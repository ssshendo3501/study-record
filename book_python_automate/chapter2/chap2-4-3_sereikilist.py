import openpyxl as excel

# 西暦と和暦のテーブル
wareki_table = [{'name': '明治', 'start': 1868, 'end': 1912}, 
                {'name': '大正', 'start': 1912, 'end': 1926}, 
                {'name': '昭和', 'start': 1926, 'end': 1989}, 
                {'name': '平成', 'start': 1989, 'end': 2019}, 
                {'name': '令和', 'start': 2019, 'end': 9999} ]


# 西暦から和暦へ変更する関数
def sereki_wareki(year):
    for w in wareki_table:
        if w['start'] <= year < w['end']:
            y = str(year - w['start'] +1) + "年"
            if y == "1年": y == "元年"
            return w['name'] + y
    return "不明"

# print(sereki_wareki(1869))

# 新規ワークブックを作成しシートを作る
book = excel.Workbook()
sheet = book.active

# シートのヘッダー部分に説明を入れる
sheet['A1'] = "西暦"
sheet['B1'] = "和暦"

# 100年分の西暦和暦対応表を作る
start_y = 1930
for i in range(100):
    # 西暦と和暦を計算
    sei = start_y + i
    wa = sereki_wareki(sei)
    # シートに設定
    sheet.cell(row=2+i, column=1, value=str(sei)+'年')
    sheet.cell(row=2+i, column=2, value=wa)
    print(sei, '=', wa)

book.save("seireki.xlsx")