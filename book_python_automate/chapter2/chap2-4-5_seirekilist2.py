import openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet['A1'] = "西暦"
sheet['B1'] = "和暦"

start_y = 1950
for i in range(80):
    sei = start_y + i
    wa = f'=TEXT({sei}/1/1, "ggge年")'
    
    sheet.cell(row=(2+i), column=1, value=str(sei)+"年")
    sheet.cell(row=(2+i), column=2, value=wa)
    
book.save("seireki2.xlsx")
    
