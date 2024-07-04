import openpyxl as xl

book = xl.Workbook()
sheet = book.active

# 横幅を指定 --- (*1)
sheet.column_dimensions['B'].width = 40
# 高さを指定 --- (*2)
sheet.row_dimensions[2].height = 40 

cell = sheet["B2"]
cell.value = "喜びにあふれた心は良い薬"

# テキスト配置の指定 --- (*3)
from openpyxl.styles.alignment import Alignment
cell.alignment = Alignment(
        horizontal='center', # 水平位置を中央に
        vertical='center') # 垂直位置を中央に

# 罫線の指定 --- (*4)
from openpyxl.styles.borders import Border, Side
cell.border = Border(
    top=Side(style='thin', color='000000'),   # 上
    right=Side(style='thin', color='000000'), # 右
    bottom=Side(style='thin', color='000000'),# 下
    left=Side(style='thin', color='000000'),  # 左
)

# フォントの指定 --- (*5)
from openpyxl.styles import Font
cell.font = Font(
    size=14,        # 文字サイズ
    bold=True,      # 太文字
    italic=True,    # イタリック
    color='FFFFFF') # 文字色

# 背景色の指定 --- (*6)
from openpyxl.styles import PatternFill
cell.fill = PatternFill(
        fill_type='solid', # べた
        fgColor='FF0000')  # 赤色

# 保存
book.save("style.xlsx")

