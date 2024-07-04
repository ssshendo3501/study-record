# プランごとに別シートに分ける

import openpyxl as excel

book = excel.load_workbook("all-customer.xlsx")
sheet = book["名簿"]

# 顧客一覧を確認してシートを分ける
for row in sheet.iter_rows(min_row=3):
    cells = [v.value for v in row]
    if cells[0] is None: break
    print(cells)
    
    # 各行のデータを変数に代入
    (name, area, plan) = cells
    
    # コピー先のシート名
    sname = plan + "プラン"
    
    # 該当するシートがあるか？
    if sname not in book.sheetnames:
        to_sheet = book.create_sheet(title=sname)
        to_sheet.append(["名前", "住所", "プラン"])
    else:
        to_sheet = book[sname]
    
    # 該当シートに顧客情報を追記
    to_sheet.append(cells)

book.save("split_sheet.xlsx")