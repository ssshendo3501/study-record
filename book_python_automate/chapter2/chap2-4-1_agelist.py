# 年齢生年対応表
# 年齢生年対応表でやっかいなのは、年が変わると全面的に更新が必要であること
# Pythonで自動更新できるようにしておけば、毎年手作業で更新しなくて良くなる


import openpyxl as excel
import datetime 

book = excel.Workbook()
sheet = book.active

# 今年の年を得る
thisyear = datetime.date.today().year

# 連続でｓ流に値を設定する
for i in range(80):
    # 設定する値を計算
    age = i  # 満年齢
    year = thisyear - i  # 生まれ年
    # セルを取得し値を設定
    age_cell = sheet.cell(row=i+1, column=1)
    age_cell.value = str(i) + "才"
    year_cell = sheet.cell(row=i+1, column=2)
    year_cell.value = str(year) + "年"
    
book.save("agelist.xlsx")