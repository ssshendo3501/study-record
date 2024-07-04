import openpyxl as excel

# テンプレートから請求書を作成
# ファイル名の指定
templete_file = 'invoice-templete.xlsx'
save_file = 'invoice01.xlsx'

# 設定データ
name = '田中一郎'
subject = '1月のご請求'
items = [# 内訳
         ['リンゴ', 5, 320],
         ['バナナ', 8, 210],
         ['メロン', 1, 1200]
         ]

# テンプレートを開く
book = excel.load_workbook('invoice-template.xlsx')
sheet = book.active

# テンプレートに名前と件名を書き込む
sheet['B4'] = name
sheet['C10'] = subject
# 内訳を連続で書き込む
total = 0
for i, it in enumerate(items):
    summary, count, price = it
    subtotal =count * price
    total += subtotal
    # シートに書き込む
    row = 15 + i
    sheet.cell(row, 2, summary)
    sheet.cell(row, 5, count)
    sheet.cell(row, 6, price)
    sheet.cell(row, 7, subtotal)
# 請求金額合計を書き込む
sheet['C11'] = total

book.save(save_file)
    