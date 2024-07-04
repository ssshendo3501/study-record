# csvファイルを読み込んで、Excelシートに書き込むプログラムを作る
# 名簿管理アプリから出力したcsvファイルを、ラベルシートっぽいフォーマットのExcelファイルに書き込む
# ラベルシートのフォーマットは1シートに10枚しかラベルカードがないので、10を超えたら改ページする必要がある

import csv
import openpyxl as excel


# 設定
in_file = 'meibo.csv'
templete_file  = 'card-template.xlsx'
save_file = 'card-meibo.xlsx'


# csvファイルを読み込む
def read_csv(fname):
    with open(fname, encoding='sjis') as f:
        reader = csv.reader(f)
        next(reader)
        return [v for v in reader]

# Excelブックを読み込みテンプレートのシートを得る
book = excel.load_workbook(templete_file)
sheet_tpl = book['Sheet']

# csvから顧客一覧を得て一人一人処理
for i, person in enumerate(read_csv(in_file)):
    # csvの一行を変数に割り振る
    name, zipno, addr = person
    # 1枚のページに10枚ずつ
    idx = i % 10
    if idx == 0:
        # テンプレートのシートコピー
        sheet = book.copy_worksheet(sheet_tpl)
        sheet.titel = 'Page' + str(idx)
    # 書き込むセル位置の決定
    row = 4 * (idx % 5) + 2
    col = 2 * (idx // 5) + 2
    # セルにデータを書き込む
    sheet.cell(row=row+0, column=col, value=name)
    sheet.cell(row=row+1, column=col, value=zipno)
    sheet.cell(row=row+2, column=col, value=addr)

# テンプレートのシートを削除して保存
book.remove(sheet_tpl)
book.save(save_file)