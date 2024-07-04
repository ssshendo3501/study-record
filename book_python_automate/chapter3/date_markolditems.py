# 条件付き書式を使わずに任意のセルに色をつける
# 古い在庫が記録されているセルがあれば、セルを赤でマーキングし強調表示するようにする

import openpyxl as excel 
from openpyxl.styles import PatternFill
import datetime


in_file = 'date-sample.xlsx'
out_file = 'date-mark-olditems.xlsx'
limit = datetime.datetime(2020, 1, 1)


# main処理
def date_mark(in_file, out_file):
    # Excelブックを開く
    book = excel.load_workbook(in_file)
    # 全シートを確認する
    for sheet in book.worksheets:
        # シート内の範囲を調べる
        for row in sheet.iter_rows(min_row=4):
            check_cell(row)
    # 保存
    book.save(out_file)
    

# 指定行をチェックする
def check_cell(row):
    # A列を得る
    date = row[0].value
    # セルが日付型でなければチェックしない
    if type(date) is not datetime.datetime:
        return
    # 古い在庫か？
    if date < limit:
        # 背景を赤色に
        red = PatternFill(fill_type='solid', fgColor='ff0000')
        # その行のセルを全て赤色に
        for cell in row:
            cell.fill = red


if __name__ == '__main__':
    date_mark(in_file, out_file)